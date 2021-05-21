import openpyxl
wb=openpyxl.load_workbook("2.xlsx")
print(wb)

sheet=(wb.active.title)
print(sheet)
actsht=wb.sheetnames
print(actsht)

data=wb["first"]["D16"].value
print(data)


data2=wb["second"]["C20"].value
print(data2)

sht2=wb["second"]
row=sht2.max_row
column=sht2.max_column
print(row)
print(column)

for i in range(1,row+1):
    for j in range (1,column+1):
        print(sht2.cell(i,j).value)

sht2.cell(row=31,column=1,value="END")

wb.save("2.xlsx")


